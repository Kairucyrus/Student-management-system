from tkinter import *
#import tkinter as tk
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib


background = "#232B2B"  #232B2B
framebg = "#EBE8DB"
framefg = "#06283D"
entrybg = "#E5D0AC"

root = Tk()
root.title("Student Registration")
root.geometry("1280x700+150+70") #1280-width, 700-height, 210-leftmargin, 100-topmargin
root.minsize(1280,700)
root.config(bg=background)

wb= pathlib.Path('Student_data.xlsx')

if wb.exists():
    pass
else:
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Registration Number"
    ws['B1'] = "First Name"
    ws['C1'] = "Second Name"
    ws['D1'] = "Surname"
    ws['E1'] = "Class"
    ws['F1'] = "Gender"
    ws['G1'] = "Date Of Birth"
    ws['H1'] = "Registration Year"
    ws['I1'] = "Registration Month"
    ws['J1'] = "Registration Day"
    ws['K1'] = "Special Needs"
    ws['L1'] = "Co-Curricular"
    ws['M1'] = "Father's Name"
    ws['N1'] = "Mother's namer"
    ws['O1'] = "Father's Occupation"
    ws['P1'] = "Mother's Occupation"

    wb.save("Student_data.xlsx")



def gender_selection():
    value=radio.get()

    if value == 1:
        gender="Male"
        print(gender)
    else:
        gender = "Female"
        
    
    return gender

def add_placeholder(entry, placeholder, color='#000'):
    entry.insert(0, placeholder)
    entry.config(fg=color)

    def on_focus_in(event):
        if entry.get() == placeholder:
            entry.delete(0, tk.END)
            entry.config(fg='black')

    def on_focus_out(event):
        if entry.get() == '':
            entry.insert(0, placeholder)
            entry.config(fg=color)

    entry.bind("<FocusIn>", on_focus_in)
    entry.bind("<FocusOut>", on_focus_out)

def showimage():
    global filename
    global img
    filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                          title="Select Image file", filetype=(("JPG FILE", "*.jpg"),
                                                                            ("PNG FILE","*.png"),
                                                                            ("ALL FILES", "*.txt")))

    img = (Image.open(filename))
    resized = img.resize((180, 180))
    tk_image = ImageTk.PhotoImage(resized)
    labl.config(image=tk_image)
    labl.image = tk_image

def registration_no():
    file = openpyxl.load_workbook("Student_data.xlsx")
    ws = file.active
    row = ws.max_row

    max_row_value = ws.cell(row=row, column=1).value

    try:
        Registration.set(max_row_value+1)

    except:
        Registration.set("1")


def clear():
    global img
    First_name.set('First Name')
    Second_name.set('Second Name')
    Surname.set('Optional')
    DOb.set('')
    classs_entry.set('Select Class')
    reg_year.set('Year')
    reg_month.set('Month')
    reg_day.set('Day')
    Curriculum.set('')
    DORr.set('')
    special_entry.set('Select')


    F_name.set('')
    M_name.set('')
    F_occupation.set('')
    M_occupation.set('')
    registration_no()
    save_button.config(state="normal")
    img1 = PhotoImage(file='Images/upload.png')
    labl.config(image=img1)
    labl.image = img1

    img = ""

def save():
    R1 = Registration.get()
    N1 = First_name.get()
    N2 = Second_name.get()
    S1 = Surname.get()
    C1 = classs_entry.get()

    try:
        G1 = gender

    except:
        messagebox.showerror("Error! Select Gender")

    d2 =DOb.get()
    RegY = reg_year.get()
    RegM = reg_month.get()
    RegD = reg_day.get()
    speciAl = special_entry.get()
    cc = Curriculum.get()
    fname = F_name.get()
    mname = M_name.get()
    F_oc = F_occupation.get()
    M_oc = M_occupation.get()

    """if N1 == "" or N2 == "" or cc== "" or C1 == "Select Class" or d2=='', or fname == "", or mname == "" or F_oc == "" or M_oc == "" or reg_day=="Day" or reg_month == "Month" or reg_year == "reg_year" or speciAl=="Select":
        messagebox("Missing Details!")
    """
    if N1 == "" or N2 == "" or cc == "" or C1 == "Select Class" or d2 == "" or fname == "" or mname == "" or F_oc == "" or M_oc == "" or reg_day == "Day" or reg_month == "Month" or reg_year == "reg_year" or speciAl == "Select":
        messagebox.showerror("Missing Details", "Please fill in all required fields!")


    else:
        file = openpyxl.load_workbook("Student_data.xlsx")
        ws = file.active
        ws.cell(column=1, row=ws.max_row+1, value=R1)
        ws.cell(column=2, row=ws.max_row, value=N1)
        ws.cell(column=3, row=ws.max_row, value=N2)
        ws.cell(column=4, row=ws.max_row, value=S1)
        ws.cell(column=5, row=ws.max_row, value=C1)
        ws.cell(column=6, row=ws.max_row, value=G1)
        ws.cell(column=7, row=ws.max_row, value=d2)
        ws.cell(column=8, row=ws.max_row, value=RegY)
        ws.cell(column=9, row=ws.max_row, value=RegM)
        ws.cell(column=10, row=ws.max_row, value=RegD)
        ws.cell(column=11, row=ws.max_row, value=speciAl)
        ws.cell(column=12,row=ws.max_row, value=cc)
        ws.cell(column=13, row=ws.max_row, value=fname)
        ws.cell(column=14, row=ws.max_row, value=mname)
        ws.cell(column=15, row=ws.max_row, value=F_oc)
        ws.cell(column=15, row=ws.max_row, value=M_oc)

        file.save(r'Student_data.xlsx')




    #print(R1)
    #print(N1)
    #print(N2)
    #print(C1)
    #print(G1)
    #print(d2)
    #print(RegY)
    #print(RegM)
    #print(RegD)
    #print(speciAl)
    #print(cc)
    #print(fname)
    #print(mname)
    #print(F_oc)
    #print(M_oc)
    
def gender_selection():
    global gender
    value = radio.get()
    if value == 1:
        gender="Male"

    else:
        gender= "Female"  









#top frames
Label(root, text = "Email: kairukairu442@gmail.com", font=("Lucida Console", 14, "bold"), width = 10, height=2, bg="#003153", fg= "#fff", padx = 25, anchor='e').pack(side=TOP, fill=X)
Label(root, text = "STUDENT REGISTRATION", font=("Yu Gothic", 14, "bold"), width = 10, height=2, fg="#fff",bg ="#00416A").pack(side=TOP, fill=X)

#search box
Search = StringVar()
Entry(root, textvariable=Search, width=15 , bd=2, font="Bodoni 17").place(x=900, y=55)
imageicon1 = PhotoImage(file = "Images/search2.png")
Srch = Button(root, text="Search", compound=LEFT, image=imageicon1, width=110, height=28, bg="#68ddfa", font="bodoni, 12 bold")

Srch.place(x=1130, y=53)

imageicon2 = PhotoImage(file="Images/uploaad.png")
update_button = Button(root, image=imageicon2, bg="#68ddfa", height=40, width=60)
update_button.place(x=110, y=53)

#registration and date
Label(root, text="Registration No:", font=("Arial Black", 11), fg = framebg, bg=background).place(x=40, y=110)
Label(root, text="Date:", font=("Arial Black", 13), fg = framebg, bg=background).place(x=650, y=110)
Registration = StringVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=35, font=("Times New Roman", 12), bg=entrybg)
reg_entry.place(x=200, y=110)
registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y, %H:%M:%S")
#print(d1)

date_entry = Entry(root, textvariable=Date, width=35, font=("Times New Roman", 12), bg=entrybg)
date_entry.place(x=750, y=110)

#student frame

obj1 = LabelFrame(root, text="Student Details", font = ("Roboto", 18, "bold"), bd=2, width=1050, fg=framefg, bg=framebg, height=268, relief=GROOVE)
obj1.place(x=30, y=150)

#names 

Label(root,text="First Name:", font =("Arial Black", 11), fg=framebg, bg=background).place(x=40, y=220)
Label(root,text="Second Name:", font =("Arial Black", 11), fg=framebg, bg=background).place(x=390, y=220)
Label(root,text="Surname:", font =("Arial Black", 11), fg=framebg, bg=background).place(x=770, y=220)
First_name = StringVar()
Second_name = StringVar()
Surname = StringVar()




First_name_entry = Entry(root, textvariable=First_name, width=22, font=("Times New Roman", 12), bg=entrybg)
First_name_entry.place(x=160, y=223)
add_placeholder(First_name_entry, "First Name")

Second_name_entry = Entry(root, textvariable=Second_name, width=22, font=("Times New Roman", 12), bg=entrybg)
Second_name_entry.place(x=530, y=223)
add_placeholder(Second_name_entry, "Optional")

Surname_entry = Entry(root, textvariable=Surname, width=22, font=("Times New Roman", 12), bg=entrybg)
Surname_entry.place(x=870, y=223)
add_placeholder(Surname_entry, "Surname")

#DOB and Gender
Label(root,text="Date Of Birth:", font =("Arial Black", 12), fg=framebg, bg=background).place(x=40, y=270)
Label(root,text="Gender:", font =("Arial Black", 12), fg=framebg, bg=background).place(x=650, y=270)
DOb = StringVar()
gender = StringVar()

DOb_entry = Entry(root, textvariable=DOb, width=27, font=("Times New Roman", 13), bg=entrybg)
DOb_entry.place(x=180, y=273)

radio = IntVar()
R1 = Radiobutton(text="Male", font =("Callibri", 13, "bold"), variable=radio, value=1, bg=entrybg, fg=framefg, command=gender_selection)
R1.place(x=770, y=270)
R2 = Radiobutton(text="Female", variable=radio , font =("Callibri", 13, "bold"), value=0, bg=entrybg, fg=framefg, command=gender_selection)
R2.place(x=840, y=270)

#class and Date of registration

Label(root,text="Class:", font =("Arial Black", 12), fg=framebg, bg=background).place(x=40, y=320)
Label(root,text="Registration Date:", font =("Arial Black", 13), fg=framebg, bg=background).place(x=260, y=320)
#Label(root,text="Month:", font =("Arial Black", 13), fg=framebg, bg=background).place(x=580, y=320)
#Label(root,text="Day:", font =("Arial Black", 13), fg=framebg, bg=background).place(x=840, y=320)
DORr = StringVar()

classs_entry = Combobox(root,values=["1","2","3","4","5","6","7","8","9","10","11","12"], font=("Robot", 11), width=10, state="r")
#Entry(root, textvariable=classs, width=30, font=("Times New Roman", 12), bg=entrybg)
classs_entry.place(x=110, y=323)
classs_entry.set("Select Class")

years = [str(year) for year in range(1980, 2025)]
months = ["January","February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
days = [str(day) for day in range(1, 32)]

reg_year = Combobox(root, values=years, state="r", width=12, font=("Arial", 15))
reg_year.set("Year")
reg_month = Combobox(root, values=months, state="r", width=12, font=("Arial", 15))
reg_month.set("Month")
reg_day = Combobox(root, values=days, state="r", width=12, font=("Arial", 18))
reg_day.set("Day")

reg_month.place(x=620, y=320)
reg_day.place(x=800, y=320)
#DORr_entry = Entry(root, textvariable=DORr, width=30, font=("Times New Roman", 12), bg=entrybg)
reg_year.place(x=450, y=320)

#extra-curricular activities
Label(root,text="Extra-Curriculum:", font =("Arial Black", 12), fg=framebg, bg=background).place(x=40, y=380)
Label(root,text="Special Needs:", font =("Arial Black", 12), fg=framebg, bg=background).place(x=650, y=380)
Curriculum = StringVar()
#special = StringVar()

Curriculum_entry = Entry(root, textvariable=Curriculum, width=30, font=("Times New Roman", 12), bg=entrybg)
Curriculum_entry.place(x=220, y=383)

special_entry = Combobox(root,values=["None","Visual Impairment","Hearing Impairment","Physically Handicapped","ADHD","Bipolar","Schizophrenia","Neurodevelopment","Autism","Dementia","Depressive"], font=("Arial", 13), width=24, state="r")
special_entry.set("Select")
#special_entry = Entry(root, textvariable=special, width=30, font=("Times New Roman", 12), bg=entrybg)
special_entry.place(x=810, y=380)

#separation
#Label(root, text = "Parent/Guardian Details", font=("Lucida Console", 13, "bold"), width = 30, height=1, fg="#fff",bg ="#141A1D", anchor='center').place(x=410, y=430)


#parent details
obj2 = LabelFrame(root, text="Parent Details", font = ("Roboto", 16, "bold"), bd=2, width=1050, fg=framefg, bg=framebg, height=220, relief=GROOVE)
obj2.place(x=30, y=450)

#parent/guardian-names
Label(root,text="Father's Name:", font =("Arial Black", 11), fg=framebg, bg=background).place(x=40, y=480)
Label(root,text="Mother's Name:", font =("Arial Black", 11), fg=framebg, bg=background).place(x=570, y=480)
F_name = StringVar()
M_name = StringVar()

F_name_entry = Entry(root, textvariable=F_name, width=25, font=("Times New Roman", 12), bg=entrybg)
F_name_entry.place(x=190, y=485)

M_name_entry = Entry(root, textvariable=M_name, width=25, font=("Times New Roman", 12), bg=entrybg)
M_name_entry.place(x=720, y=483)


Label(root,text="Father's Occupation:", font =("Arial Black", 11), fg=framebg, bg=background).place(x=40, y=543)
Label(root,text="Mother's Occupation:", font =("Arial Black", 11), fg=framebg, bg=background).place(x=570, y=543)
F_occupation = StringVar()
M_occupation = StringVar()

F_occupation_entry = Entry(root, textvariable=F_occupation, width=25, font=("Times New Roman", 12), bg=entrybg)
F_occupation_entry.place(x=240, y=546)

M_occupation_entry = Entry(root, textvariable=M_occupation, width=25, font=("Times New Roman", 12), bg=entrybg)
M_occupation_entry.place(x=770, y=546)


#parent/guardian occupation
#Label(root,text="Father's Occupation:", font =("Arial Black", 11), fg=framebg, bg=background).place(x=40, y=600)
#Label(root,text="Mother's Occupation:", font =("Arial Black", 11), fg=framebg, bg=background).place(x=650, y=600)
#F_occupation = StringVar()
#M_occupation = StringVar()
#
#F_occupation_entry = Entry(root, textvariable=F_occupation, width=25, font=("Times New Roman", 12), bg=entrybg)
#F_occupation_entry.place(x=240, y=603)
#
#M_occupation_entry = Entry(root, textvariable=M_occupation, width=25, font=("Times New Roman", 12), bg=entrybg)
#M_occupation_entry.place(x=840, y=603)


imag_frame = Frame(root, bg="black", bd=3, width=180, height=180, relief=GROOVE)
imag_frame.place(x=1090, y=150)

from PIL import Image, ImageTk

orig_imag= Image.open("Images/boyy.png")
resized = orig_imag.resize((180, 180))
tk_image = ImageTk.PhotoImage(resized)
#imag = PhotoImage(file="Images/boyy.png")
labl = Label(imag_frame, bg="black",image=tk_image)
labl.place(x=0, y=0)

upload_button = Button(root, text="Upload", width=10, height =2, font="Arial 13 bold", bg = "lightblue", command=showimage).place(x=1120, y=360)
save_button = Button(root, text="Save", width=10, height =2, font="Arial 13 bold", bg = "lightgreen", command=save).place(x=1120, y=430)
reset_button = Button(root, text="Reset", width=10, height =2, font="Arial 13 bold", bg = "cyan", command=clear).place(x=1120, y=500)
#save_button = Button(root, text="Upload", width=15, height =2, font="Arial 13 Bold", bg = "lightblue").place(x=1000, y=600)

#exit button
exit_button = Button(root, text="Exit", font= ("Times New Roman", 15, "bold"), width=7 , height=1, bg="#ff0000", fg="#000000", command = root.destroy)
exit_button.place(x=1130,y=630)

root.mainloop()